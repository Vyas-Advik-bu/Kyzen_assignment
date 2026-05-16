"""
Excel workbook builder.
All None/NaN values are normalized before writing — a single None passed to
an openpyxl chart crashes the entire phase, so we guard aggressively.
"""
import math
import os
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

from app.agent.schemas import Portfolio
from app.observability.logging import get_logger

log = get_logger(__name__)

_OUTPUT_DIR = Path("outputs")
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_ACCENT_FILL = PatternFill("solid", fgColor="D6E4F0")


_FORMULA_STARTERS = frozenset("=+-@\t\r\n")


def _safe(v: Any, default: Any = "N/A") -> Any:
    """Replace None, NaN, and inf with `default`."""
    if v is None:
        return default
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return default
    return v


def _safe_str(v: Any, default: str = "N/A") -> str:
    """Stringify + guard against Excel formula injection.
    Values starting with =, +, -, @, or control chars are space-prefixed
    so Excel treats them as plain text rather than formulas.
    """
    if v is None:
        return default
    s = str(v)
    if s and s[0] in _FORMULA_STARTERS:
        return " " + s
    return s


def _millions(v: float | None) -> str:
    if v is None or (isinstance(v, float) and (math.isnan(v) or math.isinf(v))):
        return "N/A"
    if abs(v) >= 1e9:
        return f"${v / 1e9:.2f}B"
    if abs(v) >= 1e6:
        return f"${v / 1e6:.1f}M"
    return f"${v:,.0f}"


def _pct(v: float | None) -> str:
    if v is None or (isinstance(v, float) and (math.isnan(v) or math.isinf(v))):
        return "N/A"
    return f"{v * 100:.1f}%"


def _header_row(ws, row: int, values: list[str]) -> None:
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def build_excel(portfolio: Portfolio, job_id: str) -> str:
    _OUTPUT_DIR.mkdir(exist_ok=True)
    path = _OUTPUT_DIR / f"{job_id}.xlsx"

    wb = openpyxl.Workbook()

    _build_overview(wb, portfolio)
    _build_income_statement(wb, portfolio)
    _build_key_metrics(wb, portfolio)
    _build_competitors(wb, portfolio)

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # On Windows, if the file is open in Excel it will be locked.
    # Fall back to a versioned filename rather than crashing Phase 3.
    for suffix in ["", "_v2", "_v3"]:
        candidate = _OUTPUT_DIR / f"{job_id}{suffix}.xlsx"
        try:
            wb.save(candidate)
            log.info("excel_saved", path=str(candidate), job_id=job_id)
            return str(candidate)
        except PermissionError:
            log.warning("excel_locked", path=str(candidate), trying_next=bool(suffix != "_v3"))
    raise PermissionError(f"All Excel paths for job {job_id} are locked")


def _build_overview(wb: openpyxl.Workbook, p: Portfolio) -> None:
    ws = wb.create_sheet("Overview")
    c = p.company
    fin = p.financials

    rows = [
        ("Company Name", _safe(c.name)),
        ("Ticker", _safe(c.ticker)),
        ("Type", _safe(c.type)),
        ("Exchange", _safe(c.exchange)),
        ("Sector", _safe(c.sector)),
        ("Industry", _safe(c.industry)),
        ("Country", _safe(c.country)),
        ("Website", _safe(c.website)),
        ("Founded", _safe(c.founded)),
        ("Headquarters", _safe(c.headquarters)),
        ("", ""),
        ("Employees", _safe(p.headcount.value if p.headcount else None)),
        ("Market Cap", _millions(fin.market_cap)),
        ("Revenue (TTM)", _millions(fin.revenue_ttm)),
        ("Enterprise Value", _millions(fin.enterprise_value)),
        ("P/E Ratio", _safe(fin.pe_ratio)),
        ("Gross Margin", _pct(fin.gross_margin)),
        ("Net Margin", _pct(fin.net_margin)),
        ("Revenue Growth YoY", _pct(fin.revenue_growth_yoy)),
        ("", ""),
        ("Market Position", _safe(p.market_position.value if p.market_position else None)),
        ("", ""),
        ("Analyst Summary", _safe(p.analyst_summary)),
    ]

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 55
    _header_row(ws, 1, ["Field", "Value"])
    for i, (label, value) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=bool(label))
        ws.cell(row=i, column=2, value=_safe_str(value))

    if p.key_products:
        r = len(rows) + 3
        ws.cell(row=r, column=1, value="Key Products / Services").font = Font(bold=True)
        for j, prod in enumerate(p.key_products[:10], start=r + 1):
            ws.cell(row=j, column=1, value=f"  • {prod}")

    if p.data_gaps:
        r = ws.max_row + 2
        ws.cell(row=r, column=1, value="Data Gaps").font = Font(bold=True, color="C00000")
        for j, gap in enumerate(p.data_gaps, start=r + 1):
            ws.cell(row=j, column=1, value=f"  ⚠ {gap}")


def _build_income_statement(wb: openpyxl.Workbook, p: Portfolio) -> None:
    ws = wb.create_sheet("Income Statement")
    annual = [yr for yr in p.financials.annual if yr.year]
    if not annual:
        ws.cell(row=1, column=1, value="No annual financial data available.")
        return

    annual_sorted = sorted(annual, key=lambda x: x.year)
    years = [str(y.year) for y in annual_sorted]
    fields = [
        ("Revenue", "revenue"),
        ("Gross Profit", "gross_profit"),
        ("Operating Income", "operating_income"),
        ("Net Income", "net_income"),
        ("EBITDA", "ebitda"),
        ("EPS", "eps"),
    ]

    _header_row(ws, 1, ["Metric"] + years)
    for row_i, (label, attr) in enumerate(fields, start=2):
        ws.cell(row=row_i, column=1, value=label).font = Font(bold=True)
        for col_i, yr in enumerate(annual_sorted, start=2):
            v = getattr(yr, attr)
            ws.cell(row=row_i, column=col_i, value=_safe(v, None))

    ws.column_dimensions["A"].width = 20
    for i in range(2, len(years) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 14

    # Revenue chart — only if we have valid numeric revenue data
    rev_values = [_safe(y.revenue, None) for y in annual_sorted]
    valid_rev = [v for v in rev_values if isinstance(v, (int, float))]
    if len(valid_rev) >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.title = f"{p.company.name} — Annual Revenue"
        chart.y_axis.title = "USD"
        chart.x_axis.title = "Year"
        chart.shape = 4
        chart.width = 20
        chart.height = 12

        data_ref = Reference(ws, min_col=2, max_col=1 + len(years),
                             min_row=2, max_row=2)
        cats_ref = Reference(ws, min_col=2, max_col=1 + len(years), min_row=1)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "A10")


def _build_key_metrics(wb: openpyxl.Workbook, p: Portfolio) -> None:
    ws = wb.create_sheet("Key Metrics")
    fin = p.financials
    metrics = [
        ("Market Capitalization", _millions(fin.market_cap)),
        ("Enterprise Value", _millions(fin.enterprise_value)),
        ("Revenue (TTM)", _millions(fin.revenue_ttm)),
        ("Gross Margin", _pct(fin.gross_margin)),
        ("Net Profit Margin", _pct(fin.net_margin)),
        ("Revenue Growth YoY", _pct(fin.revenue_growth_yoy)),
        ("P/E Ratio", _safe(fin.pe_ratio)),
        ("Employees", _safe(p.headcount.value if p.headcount else None)),
    ]
    _header_row(ws, 1, ["Metric", "Value"])
    for i, (label, value) in enumerate(metrics, start=2):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=_safe_str(value))
        if i % 2 == 0:
            for col in range(1, 3):
                ws.cell(row=i, column=col).fill = _ACCENT_FILL
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    if p.risks or p.opportunities:
        r = len(metrics) + 3
        if p.risks:
            ws.cell(row=r, column=1, value="Key Risks").font = Font(bold=True)
            for j, risk in enumerate(p.risks[:5], start=r + 1):
                ws.cell(row=j, column=1, value=f"• {risk}")
            r = ws.max_row + 2
        if p.opportunities:
            ws.cell(row=r, column=1, value="Opportunities").font = Font(bold=True)
            for j, opp in enumerate(p.opportunities[:5], start=r + 1):
                ws.cell(row=j, column=1, value=f"• {opp}")


def _build_competitors(wb: openpyxl.Workbook, p: Portfolio) -> None:
    ws = wb.create_sheet("Competitors")
    if not p.competitors:
        ws.cell(row=1, column=1, value="No competitor data available.")
        return
    _header_row(ws, 1, ["Company", "Ticker", "Market Cap", "Revenue (TTM)", "Summary"])
    for i, comp in enumerate(p.competitors, start=2):
        ws.cell(row=i, column=1, value=_safe_str(comp.name))
        ws.cell(row=i, column=2, value=_safe_str(comp.ticker))
        ws.cell(row=i, column=3, value=_millions(comp.market_cap))
        ws.cell(row=i, column=4, value=_millions(comp.revenue_ttm))
        ws.cell(row=i, column=5, value=_safe_str(comp.summary))
    for col, width in zip("ABCDE", [25, 10, 16, 16, 50]):
        ws.column_dimensions[col].width = width
