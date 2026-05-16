"""Excel builder tests — no Ollama or network needed."""
import math
import os
from pathlib import Path

import openpyxl
import pytest

from app.agent.schemas import (
    Portfolio, ResolvedCompany, Financials, FinancialYear,
    Competitor, Evidence, Confidence, CompanyType,
)
from app.excel.builder import build_excel, _safe, _millions, _pct


class TestHelpers:
    def test_safe_none(self):
        assert _safe(None) == "N/A"

    def test_safe_nan(self):
        assert _safe(float("nan")) == "N/A"

    def test_safe_inf(self):
        assert _safe(float("inf")) == "N/A"

    def test_safe_zero(self):
        assert _safe(0) == 0

    def test_millions_billions(self):
        assert "B" in _millions(3e9)

    def test_millions_millions(self):
        assert "M" in _millions(5e7)

    def test_millions_none(self):
        assert _millions(None) == "N/A"

    def test_pct_formats(self):
        assert _pct(0.25) == "25.0%"

    def test_pct_none(self):
        assert _pct(None) == "N/A"


class TestBuildExcel:
    @pytest.fixture
    def sample_portfolio(self) -> Portfolio:
        return Portfolio(
            company=ResolvedCompany(
                name="Acme Corp", ticker="ACME",
                type=CompanyType.PUBLIC, sector="Technology",
            ),
            financials=Financials(
                market_cap=5e9,
                revenue_ttm=1e9,
                gross_margin=0.6,
                net_margin=0.15,
                annual=[
                    FinancialYear(year=2023, revenue=1e9, net_income=1.5e8),
                    FinancialYear(year=2022, revenue=8e8, net_income=1.2e8),
                ],
            ),
            headcount=Evidence(value=10000, source="web", confidence=Confidence.HIGH),
            competitors=[
                Competitor(name="Rival Co", ticker="RIVL", market_cap=3e9),
            ],
            key_products=["Widget Pro", "Widget Lite"],
            data_gaps=["No private revenue data"],
        )

    def test_creates_file(self, sample_portfolio, tmp_path, monkeypatch):
        monkeypatch.chdir(tmp_path)
        path = build_excel(sample_portfolio, "test-job-123")
        assert Path(path).exists()

    def test_expected_sheets(self, sample_portfolio, tmp_path, monkeypatch):
        monkeypatch.chdir(tmp_path)
        path = build_excel(sample_portfolio, "test-job-456")
        wb = openpyxl.load_workbook(path)
        assert set(wb.sheetnames) == {"Overview", "Income Statement", "Key Metrics", "Competitors"}

    def test_handles_all_none_financials(self, tmp_path, monkeypatch):
        monkeypatch.chdir(tmp_path)
        p = Portfolio(company=ResolvedCompany(name="Ghost Co"))
        path = build_excel(p, "test-job-789")
        wb = openpyxl.load_workbook(path)
        ws = wb["Income Statement"]
        assert ws.cell(1, 1).value is not None  # has content even with no data

    def test_nan_values_do_not_crash(self, tmp_path, monkeypatch):
        monkeypatch.chdir(tmp_path)
        p = Portfolio(
            company=ResolvedCompany(name="NaN Corp"),
            financials=Financials(
                market_cap=float("nan"),
                revenue_ttm=None,
                annual=[FinancialYear(year=2023, revenue=None)],
            ),
        )
        path = build_excel(p, "test-job-nan")
        assert Path(path).exists()
